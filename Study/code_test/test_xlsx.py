# -*- coding: utf-8 -*-
import openpyxl

#打开文件
file_path = "C:/Users/ChunhuaGuo/Desktop/瑞福花园线上问题checklist.xlsx"
wp = openpyxl.load_workbook(file_path)

print("获取所有工作表格名：",wp.get_sheet_names())

#sheet1 = wp.get_active_sheet()
#print("获取活动的工作表名：",sheet1.title)

#获取单元格数据
#print(sheet1['A1'].value) #获取单元格A1的值
#print(sheet1.cell(row=2,column=1).value) #获取第2行第1列中单元格的值

#数据写入
sheet2 = wp.get_active_sheet()
print("获取活动的工作表名：",sheet2.title)
sheet2.cell(row=4,column=1).value = "工作总结"
print(sheet2.cell(row=4,column=1).value)
sheet2['C3'] = 'cs'
print(sheet2['C3'].value)
print("最大列数",sheet2.max_column)
print("最大行数",sheet2.max_row)

wp.save("open.xlsx")